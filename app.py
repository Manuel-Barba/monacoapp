from flask import Flask, render_template, jsonify, request, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from datetime import datetime, timedelta
import pytz
import os
import io
import cairosvg
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from mesas_config import get_mesas_config, get_layout_config, get_mesas_por_area as get_mesas_config_por_area, get_mesa_config

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///restaurant.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
migrate = Migrate(app, db)

# Configurar zona horaria del restaurante (GMT-7)
RESTAURANT_TIMEZONE = pytz.timezone('America/Phoenix')  # GMT-7 (sin horario de verano)

def get_restaurant_now():
    """Obtiene la fecha y hora actual en la zona horaria del restaurante"""
    return datetime.now(RESTAURANT_TIMEZONE)

class Mesa(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    numero = db.Column(db.Integer, nullable=False)
    capacidad = db.Column(db.Integer, nullable=False)
    estado = db.Column(db.String(20), default='disponible')  # disponible, ocupada, reservada
    ubicacion = db.Column(db.String(50))  # interior, jardin, reservados
    posicion_x = db.Column(db.Integer)  # Para posicionamiento en el layout
    posicion_y = db.Column(db.Integer)  # Para posicionamiento en el layout
    grupo_id = db.Column(db.Integer, nullable=True)  # Para agrupar mesas
    fecha = db.Column(db.Date, nullable=True)  # Fecha de ocupación (en GMT-7)

    @property
    def capacidad_total(self):
        if self.grupo_id:
            mesas_grupo = Mesa.query.filter_by(grupo_id=self.grupo_id).all()
            return sum(mesa.capacidad for mesa in mesas_grupo)
        return self.capacidad

    @property
    def mesas_grupo(self):
        if self.grupo_id:
            return Mesa.query.filter_by(grupo_id=self.grupo_id).all()
        return [self]

class Reservacion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    mesa_id = db.Column(db.Integer, db.ForeignKey('mesa.id'), nullable=False)
    hora_reservacion = db.Column(db.Time, nullable=False)
    area = db.Column(db.String(50), nullable=False)  # interior, jardin, reservados
    cantidad_personas = db.Column(db.Integer, nullable=False)
    nombre_reservador = db.Column(db.String(100), nullable=False)
    telefono = db.Column(db.String(20), nullable=True)  # Nuevo campo
    nota = db.Column(db.Text, nullable=True)  # Nuevo campo
    fecha_reservacion = db.Column(db.Date, nullable=False)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relación con la mesa
    mesa = db.relationship('Mesa', backref=db.backref('reservaciones', lazy=True))
    
    def to_dict(self):
        return {
            'id': self.id,
            'mesa_id': self.mesa_id,
            'mesa_numero': self.mesa.numero if self.mesa else None,
            'hora_reservacion': self.hora_reservacion.strftime('%H:%M') if self.hora_reservacion else None,
            'area': self.area,
            'cantidad_personas': self.cantidad_personas,
            'nombre_reservador': self.nombre_reservador,
            'telefono': self.telefono,
            'nota': self.nota,
            'fecha_reservacion': self.fecha_reservacion.strftime('%Y-%m-%d') if self.fecha_reservacion else None,
            'fecha_creacion': self.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if self.fecha_creacion else None
        }

class HistorialReservacion(db.Model):
    """Tabla para mantener un registro de las reservaciones liberadas"""
    id = db.Column(db.Integer, primary_key=True)
    reservacion_id_original = db.Column(db.Integer, nullable=False)  # ID de la reservación original
    mesa_id = db.Column(db.Integer, nullable=False)
    mesa_numero = db.Column(db.Integer, nullable=False)
    hora_reservacion = db.Column(db.Time, nullable=False)
    area = db.Column(db.String(50), nullable=False)
    cantidad_personas = db.Column(db.Integer, nullable=False)
    nombre_reservador = db.Column(db.String(100), nullable=False)
    telefono = db.Column(db.String(20), nullable=True)  # Nuevo campo
    nota = db.Column(db.Text, nullable=True)  # Nuevo campo
    fecha_reservacion = db.Column(db.Date, nullable=False)
    fecha_creacion_original = db.Column(db.DateTime, nullable=False)
    fecha_liberacion = db.Column(db.DateTime, default=datetime.utcnow)
    hora_liberacion = db.Column(db.Time, nullable=True)  # Hora exacta de liberación
    motivo_liberacion = db.Column(db.String(200), default='Liberada por el usuario')
    
    def to_dict(self):
        return {
            'id': self.id,
            'reservacion_id_original': self.reservacion_id_original,
            'mesa_id': self.mesa_id,
            'mesa_numero': self.mesa_numero,
            'hora_reservacion': self.hora_reservacion.strftime('%H:%M') if self.hora_reservacion else None,
            'area': self.area,
            'cantidad_personas': self.cantidad_personas,
            'nombre_reservador': self.nombre_reservador,
            'telefono': self.telefono,
            'nota': self.nota,
            'fecha_reservacion': self.fecha_reservacion.strftime('%Y-%m-%d') if self.fecha_reservacion else None,
            'fecha_creacion_original': self.fecha_creacion_original.strftime('%Y-%m-%d %H:%M:%S') if self.fecha_creacion_original else None,
            'fecha_liberacion': self.fecha_liberacion.strftime('%Y-%m-%d %H:%M:%S') if self.fecha_liberacion else None,
            'hora_liberacion': self.hora_liberacion.strftime('%H:%M') if self.hora_liberacion else None,
            'motivo_liberacion': self.motivo_liberacion
        }

@app.route('/')
def home():
    mesas = Mesa.query.all()
    return render_template('index.html', mesas=mesas)

@app.route('/reservaciones-futuras')
def reservaciones_futuras():
    return render_template('reservaciones_futuras.html')

@app.route('/api/fecha-actual', methods=['GET'])
def get_fecha_actual():
    """Obtiene la fecha actual en la zona horaria del restaurante"""
    fecha_actual = get_restaurant_now().date()
    return jsonify({
        'fecha': fecha_actual.strftime('%Y-%m-%d'),
        'zona_horaria': 'America/Phoenix (GMT-7)'
    })

@app.route('/api/mesas', methods=['GET'])
def get_mesas():
    """API optimizada que combina configuración estática con estados dinámicos de BD"""
    fecha_actual = get_restaurant_now().date()
    
    # Obtener configuración estática de todas las mesas
    mesas_config = get_mesas_config()
    
    # Obtener estados dinámicos de la base de datos (una sola consulta)
    mesas_db = Mesa.query.all()
    
    # Crear diccionario de estados por número de mesa para acceso rápido
    estados_mesas = {}
    for mesa_db in mesas_db:
        estados_mesas[mesa_db.numero] = {
            'id': mesa_db.id,
            'estado': mesa_db.estado,
            'grupo_id': mesa_db.grupo_id,
            'fecha': mesa_db.fecha.strftime('%Y-%m-%d') if mesa_db.fecha else None
        }
    
    # Combinar configuración estática con estados dinámicos
    mesas_data = []
    for area, mesas_area in mesas_config.items():
        for mesa_config in mesas_area:
            numero = mesa_config['numero']
            estado_mesa = estados_mesas.get(numero, {
                'id': None,
                'estado': 'disponible',
                'grupo_id': None,
                'fecha': None
            })
            
            mesa_data = {
                'id': estado_mesa['id'],
                'numero': numero,
                'capacidad': mesa_config['capacidad'],
                'estado': estado_mesa['estado'],
                'ubicacion': area,
                'posicion_x': mesa_config['posicion_x'],
                'posicion_y': mesa_config['posicion_y'],
                'grupo_id': estado_mesa['grupo_id'],
                'fecha': estado_mesa['fecha'],
                'mesas_grupo': None,  # Se calcula dinámicamente si es necesario
                'reservaciones': []  # Se carga por separado si es necesario
            }
            
            mesas_data.append(mesa_data)
    
    return jsonify(mesas_data)

@app.route('/api/mesas/<int:mesa_id>', methods=['PUT'])
def actualizar_estado_mesa(mesa_id):
    mesa = db.session.get(Mesa, mesa_id)
    data = request.get_json()
    
    if 'estado' in data:
        mesa.estado = data['estado']
        # Si la mesa se marca como ocupada, asignar la fecha enviada o la fecha actual en GMT-7
        if data['estado'] == 'ocupada':
            if 'fecha' in data and data['fecha']:
                mesa.fecha = datetime.strptime(data['fecha'], '%Y-%m-%d').date()
            else:
                mesa.fecha = get_restaurant_now().date()
        # Si la mesa se marca como disponible, limpiar la fecha
        elif data['estado'] == 'disponible':
            mesa.fecha = None
    
    if 'grupo_id' in data:
        mesa.grupo_id = data['grupo_id']
    
    db.session.commit()
    return jsonify({'mensaje': 'Estado actualizado correctamente'})

@app.route('/api/mesas/grupo', methods=['POST'])
def unir_mesas():
    data = request.get_json()
    mesa_principal_id = data.get('mesa_principal_id')
    mesa_secundaria_id = data.get('mesa_secundaria_id')
    
    mesa_principal = Mesa.query.get_or_404(mesa_principal_id)
    mesa_secundaria = Mesa.query.get_or_404(mesa_secundaria_id)
    
    # Verificar que las mesas estén en la misma ubicación
    if mesa_principal.ubicacion != mesa_secundaria.ubicacion:
        return jsonify({'error': 'Las mesas deben estar en la misma ubicación'}), 400
    
    # Si la mesa secundaria ya está en un grupo, no permitir la unión
    if mesa_secundaria.grupo_id and mesa_secundaria.grupo_id != mesa_principal_id:
        return jsonify({'error': 'La mesa secundaria ya pertenece a otro grupo'}), 400
    
    # Si la mesa principal ya está en un grupo, usar ese grupo_id
    # Si no, crear un nuevo grupo_id usando el ID de la mesa principal
    grupo_id = mesa_principal.grupo_id or mesa_principal.id
    
    # Unir las mesas
    mesa_principal.grupo_id = grupo_id
    mesa_secundaria.grupo_id = grupo_id
    
    # Si la mesa principal está ocupada, la secundaria adopta ese estado y fecha
    if mesa_principal.estado == 'ocupada':
        mesa_secundaria.estado = mesa_principal.estado
        mesa_secundaria.fecha = mesa_principal.fecha
    # Si la mesa secundaria está ocupada y la principal no, la principal adopta ese estado y fecha
    elif mesa_secundaria.estado == 'ocupada' and mesa_principal.estado == 'disponible':
        mesa_principal.estado = mesa_secundaria.estado
        mesa_principal.fecha = mesa_secundaria.fecha
    
    db.session.commit()
    
    return jsonify({
        'mensaje': 'Mesas unidas correctamente',
        'grupo_id': grupo_id,
        'capacidad_total': mesa_principal.capacidad_total
    })

@app.route('/api/mesas/grupo/<int:mesa_id>', methods=['DELETE'])
def separar_mesas(mesa_id):
    mesa = Mesa.query.get_or_404(mesa_id)
    
    if not mesa.grupo_id:
        return jsonify({'error': 'La mesa no está en un grupo'}), 400
    
    # Obtener todas las mesas del grupo
    mesas_grupo = Mesa.query.filter_by(grupo_id=mesa.grupo_id).all()
    
    # Separar todas las mesas del grupo
    for mesa_grupo in mesas_grupo:
        mesa_grupo.grupo_id = None
        # Mantener el estado y fecha actuales al separar
    
    db.session.commit()
    return jsonify({'mensaje': 'Grupo separado correctamente'})

@app.route('/api/mesas/area/<area>', methods=['GET'])
def get_mesas_por_area(area):
    """API optimizada para obtener mesas por área usando configuración estática"""
    fecha_actual = get_restaurant_now().date()
    
    # Obtener configuración estática para el área
    mesas_config_area = get_mesas_config_por_area(area)
    if not mesas_config_area:
        return jsonify([])
    
    # Obtener números de mesa para este área
    numeros_mesas_area = [mesa['numero'] for mesa in mesas_config_area]
    
    # Obtener estados dinámicos solo para las mesas de este área
    mesas_db = Mesa.query.filter(Mesa.numero.in_(numeros_mesas_area)).all()
    
    # Crear diccionario de estados por número de mesa
    estados_mesas = {}
    for mesa_db in mesas_db:
        estados_mesas[mesa_db.numero] = {
            'id': mesa_db.id,
            'estado': mesa_db.estado,
            'grupo_id': mesa_db.grupo_id,
            'fecha': mesa_db.fecha.strftime('%Y-%m-%d') if mesa_db.fecha else None
        }
    
    # Combinar configuración estática con estados dinámicos
    mesas_data = []
    for mesa_config in mesas_config_area:
        numero = mesa_config['numero']
        estado_mesa = estados_mesas.get(numero, {
            'id': None,
            'estado': 'disponible',
            'grupo_id': None,
            'fecha': None
        })
        
        mesa_data = {
            'id': estado_mesa['id'],
            'numero': numero,
            'capacidad': mesa_config['capacidad'],
            'estado': estado_mesa['estado'],
            'ubicacion': area,
            'posicion_x': mesa_config['posicion_x'],
            'posicion_y': mesa_config['posicion_y'],
            'grupo_id': estado_mesa['grupo_id'],
            'fecha': estado_mesa['fecha'],
            'mesas_grupo': None,  # Se calcula dinámicamente si es necesario
            'reservaciones': []  # Se carga por separado si es necesario
        }
        
        mesas_data.append(mesa_data)
    
    return jsonify(mesas_data)

# Endpoints para reservaciones
@app.route('/api/reservaciones', methods=['GET'])
def get_reservaciones():
    fecha = request.args.get('fecha')
    if fecha:
        try:
            fecha_parseada = datetime.strptime(fecha, '%Y-%m-%d').date()
            reservaciones = Reservacion.query.filter_by(fecha_reservacion=fecha_parseada).all()
        except ValueError:
            return jsonify({'error': 'Formato de fecha inválido'}), 400
    else:
        reservaciones = Reservacion.query.all()
    
    return jsonify([reservacion.to_dict() for reservacion in reservaciones])

@app.route('/api/reservaciones', methods=['POST'])
def crear_reservacion():
    data = request.get_json()
    
    # Validar datos requeridos
    required_fields = ['mesa_id', 'hora_reservacion', 'area', 'cantidad_personas', 'nombre_reservador', 'fecha_reservacion']
    for field in required_fields:
        if field not in data or not data[field]:
            return jsonify({'error': f'Campo requerido: {field}'}), 400
    
    # Verificar que la mesa existe
    mesa = Mesa.query.get(data['mesa_id'])
    if not mesa:
        return jsonify({'error': 'Mesa no encontrada'}), 404
    
    # Verificar que la mesa esté disponible para la fecha específica
    # Una mesa puede estar "disponible" pero tener reservaciones en otras fechas
    fecha_reservacion = datetime.strptime(data['fecha_reservacion'], '%Y-%m-%d').date()
    
    # Si la mesa está ocupada, no se puede reservar
    if mesa.estado == 'ocupada':
        return jsonify({'error': 'La mesa está ocupada y no se puede reservar'}), 400
    
    # Si la mesa está reservada, verificar si es para la misma fecha
    if mesa.estado == 'reservada' and mesa.fecha == fecha_reservacion:
        return jsonify({'error': 'La mesa ya está reservada para esta fecha'}), 400
    
    # Verificar que no haya conflicto de horarios para la misma mesa y fecha
    hora_reservacion = datetime.strptime(data['hora_reservacion'], '%H:%M').time()
    
    # Buscar reservaciones existentes para la misma mesa y fecha
    reservaciones_existentes = Reservacion.query.filter_by(
        mesa_id=data['mesa_id'],
        fecha_reservacion=fecha_reservacion
    ).all()
    
    # Verificar conflictos de horario (asumiendo que una reservación dura 2 horas)
    for reservacion_existente in reservaciones_existentes:
        hora_existente = reservacion_existente.hora_reservacion
        # Calcular si hay solapamiento de horarios
        if abs((hora_reservacion.hour * 60 + hora_reservacion.minute) - 
               (hora_existente.hour * 60 + hora_existente.minute)) < 120:
            return jsonify({'error': 'Ya existe una reservación para esta mesa en ese horario'}), 400
    
    try:
        # Crear la reservación
        nueva_reservacion = Reservacion(
            mesa_id=data['mesa_id'],
            hora_reservacion=hora_reservacion,
            area=data['area'],
            cantidad_personas=data['cantidad_personas'],
            nombre_reservador=data['nombre_reservador'],
            telefono=data.get('telefono'),  # Campo opcional
            nota=data.get('nota'),  # Campo opcional
            fecha_reservacion=fecha_reservacion
        )
        
        # Cambiar el estado de la mesa solo si la reservación es para hoy
        fecha_actual = get_restaurant_now().date()
        if fecha_reservacion == fecha_actual:
            # Si la reservación es para hoy, marcar como reservada
            mesa.estado = 'reservada'
            mesa.fecha = fecha_reservacion
        else:
            # Si la reservación es para una fecha futura, mantener como disponible
            # La mesa se marcará como reservada automáticamente cuando llegue el día
            mesa.estado = 'disponible'
            mesa.fecha = None
        
        db.session.add(nueva_reservacion)
        db.session.commit()
        
        return jsonify({
            'mensaje': 'Reservación creada exitosamente',
            'reservacion': nueva_reservacion.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al crear la reservación: {str(e)}'}), 500

@app.route('/api/reservaciones/<int:reservacion_id>', methods=['DELETE'])
def eliminar_reservacion(reservacion_id):
    reservacion = Reservacion.query.get_or_404(reservacion_id)
    
    try:
        # Cambiar el estado de la mesa de vuelta a disponible
        mesa = Mesa.query.get(reservacion.mesa_id)
        if mesa:
            mesa.estado = 'disponible'
            mesa.fecha = None
        
        db.session.delete(reservacion)
        db.session.commit()
        
        return jsonify({'mensaje': 'Reservación eliminada exitosamente'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al eliminar la reservación: {str(e)}'}), 500

@app.route('/api/reservaciones/<int:reservacion_id>/liberar', methods=['POST'])
def liberar_reservacion(reservacion_id):
    """Libera una reservación moviéndola al historial y liberando la mesa"""
    reservacion = Reservacion.query.get_or_404(reservacion_id)
    
    try:
        # Obtener hora actual en zona horaria del restaurante
        hora_actual = get_restaurant_now().time()
        
        # Crear registro en el historial antes de eliminar la reservación
        historial = HistorialReservacion(
            reservacion_id_original=reservacion.id,
            mesa_id=reservacion.mesa_id,
            mesa_numero=reservacion.mesa.numero if reservacion.mesa else 0,
            hora_reservacion=reservacion.hora_reservacion,
            area=reservacion.area,
            cantidad_personas=reservacion.cantidad_personas,
            nombre_reservador=reservacion.nombre_reservador,
            telefono=reservacion.telefono,
            nota=reservacion.nota,
            fecha_reservacion=reservacion.fecha_reservacion,
            fecha_creacion_original=reservacion.fecha_creacion,
            hora_liberacion=hora_actual,
            motivo_liberacion='Liberada manualmente por el usuario'
        )
        
        # Cambiar el estado de la mesa de vuelta a disponible
        mesa = Mesa.query.get(reservacion.mesa_id)
        if mesa:
            mesa.estado = 'disponible'
            mesa.fecha = None
        
        # Agregar al historial y eliminar la reservación original
        db.session.add(historial)
        db.session.delete(reservacion)
        db.session.commit()
        
        return jsonify({
            'mensaje': 'Reservación liberada exitosamente',
            'historial': historial.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al liberar la reservación: {str(e)}'}), 500

@app.route('/api/reservaciones/mesa/<int:mesa_id>', methods=['GET'])
def get_reservaciones_mesa(mesa_id):
    reservaciones = Reservacion.query.filter_by(mesa_id=mesa_id).all()
    return jsonify([reservacion.to_dict() for reservacion in reservaciones])

@app.route('/api/mesas/especifica/<int:mesa_id>', methods=['GET'])
def get_mesa_especifica(mesa_id):
    """API optimizada para obtener una mesa específica por ID"""
    try:
        # Obtener la mesa de la base de datos
        mesa_db = Mesa.query.get(mesa_id)
        if not mesa_db:
            return jsonify({'error': 'Mesa no encontrada'}), 404
        
        # Obtener configuración estática de la mesa
        mesa_config = get_mesa_config(mesa_db.numero)
        if not mesa_config:
            return jsonify({'error': 'Configuración de mesa no encontrada'}), 404
        
        # Combinar datos estáticos con dinámicos
        mesa_data = {
            'id': mesa_db.id,
            'numero': mesa_db.numero,
            'capacidad': mesa_config['capacidad'],
            'estado': mesa_db.estado,
            'ubicacion': mesa_config['ubicacion'],
            'posicion_x': mesa_config['posicion_x'],
            'posicion_y': mesa_config['posicion_y'],
            'grupo_id': mesa_db.grupo_id,
            'fecha': mesa_db.fecha.strftime('%Y-%m-%d') if mesa_db.fecha else None
        }
        
        return jsonify(mesa_data)
        
    except Exception as e:
        return jsonify({'error': f'Error al obtener mesa: {str(e)}'}), 500

def actualizar_mesas_reservadas():
    """Función para actualizar automáticamente el estado de las mesas cuando llega el día de la reservación"""
    try:
        fecha_actual = get_restaurant_now().date()
        
        # Buscar reservaciones para hoy que tengan mesas en estado disponible
        reservaciones_hoy = Reservacion.query.filter(
            Reservacion.fecha_reservacion == fecha_actual
        ).all()
        
        actualizadas = 0
        for reservacion in reservaciones_hoy:
            mesa = Mesa.query.get(reservacion.mesa_id)
            if mesa and mesa.estado == 'disponible':
                # Cambiar el estado de la mesa a reservada
                mesa.estado = 'reservada'
                mesa.fecha = fecha_actual
                actualizadas += 1
        
        if actualizadas > 0:
            db.session.commit()
            return {
                'mensaje': f'Se actualizaron {actualizadas} mesas a estado reservado',
                'actualizadas': actualizadas
            }
        
        return None  # No devolver mensaje si no hay cambios
        
    except Exception as e:
        db.session.rollback()
        return {'error': f'Error al actualizar mesas reservadas: {str(e)}'}

def limpiar_reservaciones_pasadas():
    """Función para limpiar automáticamente las reservaciones pasadas"""
    try:
        fecha_actual = get_restaurant_now().date()
        hora_actual = get_restaurant_now().time()
        
        # Buscar reservaciones de días pasados
        reservaciones_pasadas = Reservacion.query.filter(
            Reservacion.fecha_reservacion < fecha_actual
        ).all()
        
        if not reservaciones_pasadas:
            return None  # No devolver mensaje
        
        liberadas = 0
        for reservacion in reservaciones_pasadas:
            # Crear registro en el historial
            historial = HistorialReservacion(
                reservacion_id_original=reservacion.id,
                mesa_id=reservacion.mesa_id,
                mesa_numero=reservacion.mesa.numero if reservacion.mesa else 0,
                hora_reservacion=reservacion.hora_reservacion,
                area=reservacion.area,
                cantidad_personas=reservacion.cantidad_personas,
                nombre_reservador=reservacion.nombre_reservador,
                telefono=reservacion.telefono,
                nota=reservacion.nota,
                fecha_reservacion=reservacion.fecha_reservacion,
                fecha_creacion_original=reservacion.fecha_creacion,
                hora_liberacion=hora_actual,
                motivo_liberacion='Liberación automática por fecha pasada'
            )
            
            # Liberar la mesa
            mesa = Mesa.query.get(reservacion.mesa_id)
            if mesa:
                mesa.estado = 'disponible'
                mesa.fecha = None
            
            # Agregar al historial y eliminar la reservación
            db.session.add(historial)
            db.session.delete(reservacion)
            liberadas += 1
        
        db.session.commit()
        return {
            'mensaje': f'Se liberaron {liberadas} mesas automáticamente',
            'liberadas': liberadas
        }
        
    except Exception as e:
        db.session.rollback()
        return {'error': f'Error al limpiar reservaciones: {str(e)}'}

@app.route('/api/limpiar-reservaciones-pasadas', methods=['POST'])
def api_limpiar_reservaciones_pasadas():
    """Endpoint para limpiar reservaciones pasadas manualmente"""
    resultado = limpiar_reservaciones_pasadas()
    
    if resultado is None:
        return ('', 204)  # No Content, sin mensaje
    if 'error' in resultado:
        return jsonify(resultado), 500
    
    return jsonify(resultado)

@app.route('/api/actualizar-mesas-reservadas', methods=['POST'])
def api_actualizar_mesas_reservadas():
    """Endpoint para actualizar mesas reservadas manualmente"""
    resultado = actualizar_mesas_reservadas()
    
    if resultado is None:
        return ('', 204)  # No Content, sin mensaje
    if 'error' in resultado:
        return jsonify(resultado), 500
    
    return jsonify(resultado)

@app.route('/api/actualizar-estado-mesas', methods=['POST'])
def api_actualizar_estado_mesas():
    """Endpoint para actualizar el estado de las mesas (limpiar pasadas y activar reservadas)"""
    # Primero limpiar reservaciones pasadas
    resultado_limpieza = limpiar_reservaciones_pasadas()
    
    # Luego actualizar mesas reservadas
    resultado_actualizacion = actualizar_mesas_reservadas()
    
    # Combinar resultados
    mensajes = []
    if resultado_limpieza and 'mensaje' in resultado_limpieza:
        mensajes.append(resultado_limpieza['mensaje'])
    if resultado_actualizacion and 'mensaje' in resultado_actualizacion:
        mensajes.append(resultado_actualizacion['mensaje'])
    
    if not mensajes:
        return ('', 204)  # No Content, sin mensaje
    
    return jsonify({
        'mensaje': ' | '.join(mensajes),
        'limpieza': resultado_limpieza,
        'actualizacion': resultado_actualizacion
    })

# ===== RUTAS DE EXPORTACIÓN =====

@app.route('/api/exportar-reservaciones', methods=['POST'])
def exportar_reservaciones():
    """Endpoint para exportar reservaciones en PDF o Excel"""
    try:
        data = request.get_json()
        formato = data.get('formato', 'pdf')  # pdf o excel
        fecha_inicio = datetime.strptime(data.get('fecha_inicio'), '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(data.get('fecha_fin'), '%Y-%m-%d').date()
        
        # Obtener reservaciones en el rango de fechas
        reservaciones = Reservacion.query.filter(
            Reservacion.fecha_reservacion >= fecha_inicio,
            Reservacion.fecha_reservacion <= fecha_fin
        ).order_by(Reservacion.fecha_reservacion, Reservacion.hora_reservacion).all()
        
        if formato == 'pdf':
            return generar_pdf_reservaciones(reservaciones, fecha_inicio, fecha_fin)
        elif formato == 'excel':
            return generar_excel_reservaciones(reservaciones, fecha_inicio, fecha_fin)
        else:
            return jsonify({'error': 'Formato no válido'}), 400
            
    except Exception as e:
        return jsonify({'error': f'Error al exportar: {str(e)}'}), 500

def generar_pdf_reservaciones(reservaciones, fecha_inicio, fecha_fin):
    """Genera un PDF con las reservaciones"""
    try:
        # Crear buffer para el PDF
        buffer = io.BytesIO()
        
        # Crear documento PDF en orientación horizontal
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=0.5*inch, rightMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#2C3E50')  # Azul moderno
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#7F8C8D')  # Gris moderno
        )
        
        # Agregar logo PNG en la esquina superior izquierda
        logo_path = os.path.join(app.static_folder, 'images', 'logopn.png')
        if os.path.exists(logo_path):
            try:
                logo_img = Image(logo_path, width=2*inch, height=2*inch)
                logo_img.hAlign = 'LEFT'
                story.append(logo_img)
                story.append(Spacer(1, 1))
            except Exception as e:
                # Si falla la carga del PNG, usar texto como fallback
                logo_text = Paragraph("MÓNACO BAR & GRILL", ParagraphStyle(
                    'Logo',
                    parent=styles['Heading2'],
                    fontSize=18,
                    alignment=TA_LEFT,
                    textColor=colors.HexColor('#2C3E50'),
                    spaceAfter=1
                ))
                story.append(logo_text)
        
        # Título principal
        story.append(Paragraph("Reporte de Reservaciones", title_style))
        story.append(Spacer(1, 5))
        
        # Información del rango de fechas
        fecha_texto = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} a {fecha_fin.strftime('%d/%m/%Y')}"
        story.append(Paragraph(fecha_texto, subtitle_style))
        story.append(Spacer(1, 5))
        
        # Obtener datos del historial para cálculos adicionales
        historial_reservaciones = HistorialReservacion.query.filter(
            HistorialReservacion.fecha_reservacion >= fecha_inicio,
            HistorialReservacion.fecha_reservacion <= fecha_fin
        ).all()
        
        # Combinar reservaciones activas y del historial para mostrar todas
        todas_reservaciones = []
        
        # Agregar reservaciones activas
        for reservacion in reservaciones:
            todas_reservaciones.append({
                'tipo': 'activa',
                'reservacion': reservacion,
                'historial': None
            })
        
        # Agregar reservaciones del historial
        for historial in historial_reservaciones:
            todas_reservaciones.append({
                'tipo': 'historial',
                'reservacion': None,
                'historial': historial
            })
        
        # ORDENAR todas_reservaciones por fecha y hora (ascendente)
        def get_fecha_hora(item):
            if item['tipo'] == 'activa':
                return (item['reservacion'].fecha_reservacion, item['reservacion'].hora_reservacion)
            else:
                return (item['historial'].fecha_reservacion, item['historial'].hora_reservacion)
        todas_reservaciones.sort(key=get_fecha_hora)

        # Calcular estadísticas básicas (incluyendo todas las reservaciones)
        total_reservaciones = len(todas_reservaciones)
        total_personas = 0
        for item in todas_reservaciones:
            if item['tipo'] == 'activa':
                total_personas += item['reservacion'].cantidad_personas
            else:
                total_personas += item['historial'].cantidad_personas
        
        promedio_personas = total_personas / total_reservaciones if total_reservaciones > 0 else 0
        
        # Calcular tiempo promedio de estancia
        tiempos_estancia = []
        for historial in historial_reservaciones:
            if historial.hora_liberacion and historial.hora_reservacion:
                # Calcular diferencia en minutos
                hora_inicio = historial.hora_reservacion
                hora_fin = historial.hora_liberacion
                
                # Convertir a minutos para facilitar el cálculo
                inicio_minutos = hora_inicio.hour * 60 + hora_inicio.minute
                fin_minutos = hora_fin.hour * 60 + hora_fin.minute
                
                # Si la hora de fin es menor que la de inicio, asumir que es del día siguiente
                if fin_minutos < inicio_minutos:
                    fin_minutos += 24 * 60  # Agregar 24 horas
                
                duracion_minutos = fin_minutos - inicio_minutos
                if duracion_minutos > 0:
                    tiempos_estancia.append(duracion_minutos)
        
        tiempo_promedio_estancia = sum(tiempos_estancia) / len(tiempos_estancia) if tiempos_estancia else 0
        
        # Calcular factor de ocupación por área
        ocupacion_por_area = {}
        for item in todas_reservaciones:
            if item['tipo'] == 'activa':
                area = item['reservacion'].area
                cantidad = item['reservacion'].cantidad_personas
            else:
                area = item['historial'].area
                cantidad = item['historial'].cantidad_personas
            
            if area not in ocupacion_por_area:
                ocupacion_por_area[area] = {'reservaciones': 0, 'personas': 0}
            ocupacion_por_area[area]['reservaciones'] += 1
            ocupacion_por_area[area]['personas'] += cantidad
        
        # Estadísticas básicas como lista compacta
        stats_text = f"""
        <b>Resumen del Período:</b><br/>
        • Total de Reservaciones: {total_reservaciones}<br/>
        • Total de Personas: {total_personas}<br/>
        • Promedio de personas por Reservación: {promedio_personas:.1f}<br/>
        • Tiempo Promedio de Estancia: {tiempo_promedio_estancia:.0f} min{' ' if tiempo_promedio_estancia > 0 else 'N/A'}
        """
        
        stats_paragraph = Paragraph(stats_text, ParagraphStyle(
            'Stats',
            parent=styles['Normal'],
            fontSize=11,
            spaceAfter=20,
            textColor=colors.HexColor('#2C3E50'),
            leftIndent=0
        ))
        
        story.append(stats_paragraph)
        story.append(Spacer(1, 20))
        
        # Tabla de reservaciones con hora de salida (todas las reservaciones)
        if todas_reservaciones:
            # Encabezados
            headers = ['Fecha', 'Hora', 'Cliente', 'Mesa', 'Área', 'Personas', 'Teléfono', 'Nota', 'Hora Salida', 'Estado']
            data = [headers]
            
            # Datos de todas las reservaciones
            for item in todas_reservaciones:
                if item['tipo'] == 'activa':
                    reservacion = item['reservacion']
                    fecha_formateada = reservacion.fecha_reservacion.strftime('%d/%m/%Y')
                    hora_formateada = reservacion.hora_reservacion.strftime('%H:%M')
                    mesa_numero = reservacion.mesa.numero if reservacion.mesa else 'N/A'
                    telefono = reservacion.telefono if reservacion.telefono else '-'
                    nota = reservacion.nota if reservacion.nota else '-'
                    cantidad_personas = reservacion.cantidad_personas
                    area = reservacion.area.capitalize()
                    nombre = reservacion.nombre_reservador
                    
                    # Para reservaciones activas
                    hora_salida = 'N/A'
                    estado = 'Activa'
                    
                else:
                    historial = item['historial']
                    fecha_formateada = historial.fecha_reservacion.strftime('%d/%m/%Y')
                    hora_formateada = historial.hora_reservacion.strftime('%H:%M')
                    mesa_numero = historial.mesa_numero
                    telefono = historial.telefono if historial.telefono else '-'
                    nota = historial.nota if historial.nota else '-'
                    cantidad_personas = historial.cantidad_personas
                    area = historial.area.capitalize()
                    nombre = historial.nombre_reservador
                    
                    # Para reservaciones del historial
                    if historial.hora_liberacion:
                        hora_salida = historial.hora_liberacion.strftime('%H:%M')
                        if historial.motivo_liberacion == 'Liberada manualmente por el usuario':
                            estado = 'Liberada'
                        elif historial.motivo_liberacion == 'Liberación automática por fecha pasada':
                            estado = 'No se registró hora de salida'
                        else:
                            estado = historial.motivo_liberacion
                    else:
                        hora_salida = 'N/A'
                        estado = 'Sin hora de salida'
                
                data.append([
                    fecha_formateada,
                    hora_formateada,
                    nombre,
                    str(mesa_numero),
                    area,
                    str(cantidad_personas),
                    telefono,
                    nota,
                    hora_salida,
                    estado
                ])
            
            # Crear tabla más ancha que ocupe todo el ancho de la hoja
            # En orientación horizontal A4, el ancho disponible es aproximadamente 11.7 pulgadas
            # Restamos los márgenes (1 pulgada total) = 10.7 pulgadas disponibles
            ancho_disponible = 10.7 * inch
            anchos_columnas = [
                0.8*inch,   # Fecha
                0.6*inch,   # Hora
                1.5*inch,   # Cliente
                0.5*inch,   # Mesa
                0.7*inch,   # Área
                0.5*inch,   # Personas
                1.0*inch,   # Teléfono
                1.2*inch,   # Nota
                0.8*inch,   # Hora Salida
                1.5*inch    # Estado
            ]
            
            table = Table(data, colWidths=anchos_columnas)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),  # Azul oscuro moderno
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FA')),  # Gris muy claro
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#DEE2E6')),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F8F9FA'), colors.HexColor('#FFFFFF')])  # Filas alternadas
            ]))
            
            story.append(table)
        else:
            story.append(Paragraph("No hay reservaciones en el período seleccionado", styles['Normal']))
        
        # Tabla de información adicional
        story.append(Spacer(1, 30))
        story.append(Paragraph("Información Adicional", ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=15,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#2C3E50')
        )))
        
        # Información adicional
        info_adicional = []
        
        # Ocupación por área
        for area, datos in ocupacion_por_area.items():
            promedio_area = datos['personas'] / datos['reservaciones'] if datos['reservaciones'] > 0 else 0
            info_adicional.append([
                f"Área {area.capitalize()}",
                f"{datos['reservaciones']} reservaciones",
                f"{datos['personas']} personas",
                f"{promedio_area:.1f} prom/persona"
            ])
        
        # Horario más popular
        horarios = {}
        for item in todas_reservaciones:
            if item['tipo'] == 'activa':
                hora = item['reservacion'].hora_reservacion.strftime('%H:%M')
            else:
                hora = item['historial'].hora_reservacion.strftime('%H:%M')
            horarios[hora] = horarios.get(hora, 0) + 1
        
        if horarios:
            hora_mas_popular = max(horarios, key=horarios.get)
            info_adicional.append([
                "Horario Más Popular",
                hora_mas_popular,
                f"{horarios[hora_mas_popular]} reservaciones",
                "Hora con más demanda"
            ])
        
        # Crear tabla de información adicional con colores modernos y más ancha
        if info_adicional:
            info_headers = ['Métrica', 'Valor', 'Detalle', 'Descripción']
            info_data = [info_headers] + info_adicional
            
            # Anchos de columnas para la tabla de información adicional
            info_anchos_columnas = [
                2.5*inch,   # Métrica
                2.0*inch,   # Valor
                2.0*inch,   # Detalle
                3.2*inch    # Descripción
            ]
            
            info_table = Table(info_data, colWidths=info_anchos_columnas)
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27AE60')),  # Verde moderno
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#E8F5E8')),  # Verde claro
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#BDC3C7')),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#E8F5E8'), colors.HexColor('#F8F9FA')])  # Filas alternadas
            ]))
            
            story.append(info_table)
        
        # Generar PDF
        doc.build(story)
        buffer.seek(0)
        
        # Nombre del archivo
        nombre_archivo = f"reservaciones_{fecha_inicio.strftime('%Y%m%d')}_{fecha_fin.strftime('%Y%m%d')}.pdf"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        return jsonify({'error': f'Error al generar PDF: {str(e)}'}), 500

def generar_excel_reservaciones(reservaciones, fecha_inicio, fecha_fin):
    """Genera un archivo Excel con las reservaciones"""
    try:
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reservaciones"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        data_font = Font(size=11)
        data_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:J1')
        ws['A1'] = "REPORTE DE RESERVACIONES - MÓNACO BAR & GRILL"
        ws['A1'].font = Font(bold=True, size=16, color="366092")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Período
        ws.merge_cells('A2:J2')
        ws['A2'] = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Obtener datos del historial para cálculos adicionales
        historial_reservaciones = HistorialReservacion.query.filter(
            HistorialReservacion.fecha_reservacion >= fecha_inicio,
            HistorialReservacion.fecha_reservacion <= fecha_fin
        ).all()
        
        # Combinar reservaciones activas y del historial para mostrar todas
        todas_reservaciones = []
        
        # Agregar reservaciones activas
        for reservacion in reservaciones:
            todas_reservaciones.append({
                'tipo': 'activa',
                'reservacion': reservacion,
                'historial': None
            })
        
        # Agregar reservaciones del historial
        for historial in historial_reservaciones:
            todas_reservaciones.append({
                'tipo': 'historial',
                'reservacion': None,
                'historial': historial
            })
        
        # ORDENAR todas_reservaciones por fecha y hora (ascendente)
        def get_fecha_hora(item):
            if item['tipo'] == 'activa':
                return (item['reservacion'].fecha_reservacion, item['reservacion'].hora_reservacion)
            else:
                return (item['historial'].fecha_reservacion, item['historial'].hora_reservacion)
        todas_reservaciones.sort(key=get_fecha_hora)

        # Calcular estadísticas básicas (incluyendo todas las reservaciones)
        total_reservaciones = len(todas_reservaciones)
        total_personas = 0
        for item in todas_reservaciones:
            if item['tipo'] == 'activa':
                total_personas += item['reservacion'].cantidad_personas
            else:
                total_personas += item['historial'].cantidad_personas
        
        promedio_personas = total_personas / total_reservaciones if total_reservaciones > 0 else 0
        
        # Calcular tiempo promedio de estancia
        tiempos_estancia = []
        for historial in historial_reservaciones:
            if historial.hora_liberacion and historial.hora_reservacion:
                # Calcular diferencia en minutos
                hora_inicio = historial.hora_reservacion
                hora_fin = historial.hora_liberacion
                
                # Convertir a minutos para facilitar el cálculo
                inicio_minutos = hora_inicio.hour * 60 + hora_inicio.minute
                fin_minutos = hora_fin.hour * 60 + hora_fin.minute
                
                # Si la hora de fin es menor que la de inicio, asumir que es del día siguiente
                if fin_minutos < inicio_minutos:
                    fin_minutos += 24 * 60  # Agregar 24 horas
                
                duracion_minutos = fin_minutos - inicio_minutos
                if duracion_minutos > 0:
                    tiempos_estancia.append(duracion_minutos)
        
        tiempo_promedio_estancia = sum(tiempos_estancia) / len(tiempos_estancia) if tiempos_estancia else 0
        
        # Estadísticas básicas
        ws.merge_cells('A4:B4')
        ws['A4'] = f"Total Reservaciones: {total_reservaciones}"
        ws['A4'].font = Font(bold=True)
        
        ws.merge_cells('C4:D4')
        ws['C4'] = f"Total Personas: {total_personas}"
        ws['C4'].font = Font(bold=True)
        
        ws.merge_cells('E4:F4')
        ws['E4'] = f"Promedio: {promedio_personas:.1f}"
        ws['E4'].font = Font(bold=True)
        
        ws.merge_cells('G4:H4')
        ws['G4'] = f"Tiempo Promedio: {tiempo_promedio_estancia:.0f} min" if tiempo_promedio_estancia > 0 else "Tiempo Promedio: N/A"
        ws['G4'].font = Font(bold=True)
        
        # Encabezados (fila 6)
        headers = ['Fecha', 'Hora', 'Cliente', 'Mesa', 'Área', 'Personas', 'Teléfono', 'Nota', 'Hora Salida', 'Estado']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos de todas las reservaciones
        for row, item in enumerate(todas_reservaciones, 7):
            if item['tipo'] == 'activa':
                reservacion = item['reservacion']
                fecha_formateada = reservacion.fecha_reservacion.strftime('%d/%m/%Y')
                hora_formateada = reservacion.hora_reservacion.strftime('%H:%M')
                mesa_numero = reservacion.mesa.numero if reservacion.mesa else 'N/A'
                telefono = reservacion.telefono if reservacion.telefono else '-'
                nota = reservacion.nota if reservacion.nota else '-'
                cantidad_personas = reservacion.cantidad_personas
                area = reservacion.area.capitalize()
                nombre = reservacion.nombre_reservador
                
                # Para reservaciones activas
                hora_salida = 'N/A'
                estado = 'Activa'
                
            else:
                historial = item['historial']
                fecha_formateada = historial.fecha_reservacion.strftime('%d/%m/%Y')
                hora_formateada = historial.hora_reservacion.strftime('%H:%M')
                mesa_numero = historial.mesa_numero
                telefono = historial.telefono if historial.telefono else '-'
                nota = historial.nota if historial.nota else '-'
                cantidad_personas = historial.cantidad_personas
                area = historial.area.capitalize()
                nombre = historial.nombre_reservador
                
                # Para reservaciones del historial
                if historial.hora_liberacion:
                    hora_salida = historial.hora_liberacion.strftime('%H:%M')
                    if historial.motivo_liberacion == 'Liberada manualmente por el usuario':
                        estado = 'Liberada'
                    elif historial.motivo_liberacion == 'Liberación automática por fecha pasada':
                        estado = 'No se registró hora de salida'
                    else:
                        estado = historial.motivo_liberacion
                else:
                    hora_salida = 'N/A'
                    estado = 'Sin hora de salida'
            
            data = [
                fecha_formateada,
                hora_formateada,
                nombre,
                mesa_numero,
                area,
                cantidad_personas,
                telefono,
                nota,
                hora_salida,
                estado
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
        
        # Ajustar ancho de columnas
        column_widths = [12, 8, 25, 8, 12, 10, 15, 30, 12, 25]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Crear hoja de información adicional
        ws_info = wb.create_sheet("Información Adicional")
        
        # Calcular factor de ocupación por área
        ocupacion_por_area = {}
        for item in todas_reservaciones:
            if item['tipo'] == 'activa':
                area = item['reservacion'].area
                cantidad = item['reservacion'].cantidad_personas
            else:
                area = item['historial'].area
                cantidad = item['historial'].cantidad_personas
            
            if area not in ocupacion_por_area:
                ocupacion_por_area[area] = {'reservaciones': 0, 'personas': 0}
            ocupacion_por_area[area]['reservaciones'] += 1
            ocupacion_por_area[area]['personas'] += cantidad
        
        # Título de la hoja de información
        ws_info.merge_cells('A1:D1')
        ws_info['A1'] = "INFORMACIÓN ADICIONAL - MÓNACO BAR & GRILL"
        ws_info['A1'].font = Font(bold=True, size=16, color="366092")
        ws_info['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados de información adicional
        info_headers = ['Métrica', 'Valor', 'Detalle', 'Descripción']
        for col, header in enumerate(info_headers, 1):
            cell = ws_info.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Información adicional
        row = 4
        
        # Ocupación por área
        for area, datos in ocupacion_por_area.items():
            promedio_area = datos['personas'] / datos['reservaciones'] if datos['reservaciones'] > 0 else 0
            ws_info.cell(row=row, column=1, value=f"Área {area.capitalize()}")
            ws_info.cell(row=row, column=2, value=f"{datos['reservaciones']} reservaciones")
            ws_info.cell(row=row, column=3, value=f"{datos['personas']} personas")
            ws_info.cell(row=row, column=4, value=f"{promedio_area:.1f} prom/persona")
            row += 1
        
        # Factor de eficiencia
        reservaciones_completadas = len(historial_reservaciones)
        factor_eficiencia = (reservaciones_completadas / total_reservaciones * 100) if total_reservaciones > 0 else 0
        ws_info.cell(row=row, column=1, value="Factor de Eficiencia")
        ws_info.cell(row=row, column=2, value=f"{reservaciones_completadas}/{total_reservaciones}")
        ws_info.cell(row=row, column=3, value=f"{factor_eficiencia:.1f}% completadas")
        ws_info.cell(row=row, column=4, value="Reservaciones finalizadas")
        row += 1
        
        # Horario más popular
        horarios = {}
        for item in todas_reservaciones:
            if item['tipo'] == 'activa':
                hora = item['reservacion'].hora_reservacion.strftime('%H:%M')
            else:
                hora = item['historial'].hora_reservacion.strftime('%H:%M')
            horarios[hora] = horarios.get(hora, 0) + 1
        
        if horarios:
            hora_mas_popular = max(horarios, key=horarios.get)
            ws_info.cell(row=row, column=1, value="Horario Más Popular")
            ws_info.cell(row=row, column=2, value=hora_mas_popular)
            ws_info.cell(row=row, column=3, value=f"{horarios[hora_mas_popular]} reservaciones")
            ws_info.cell(row=row, column=4, value="Hora con más demanda")
            row += 1
        
        # Aplicar estilos a la información adicional
        for r in range(4, row):
            for c in range(1, 5):
                cell = ws_info.cell(row=r, column=c)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
        
        # Ajustar ancho de columnas en la hoja de información
        info_column_widths = [25, 20, 20, 30]
        for col, width in enumerate(info_column_widths, 1):
            ws_info.column_dimensions[get_column_letter(col)].width = width
        
        # Crear buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Nombre del archivo
        nombre_archivo = f"reservaciones_{fecha_inicio.strftime('%Y%m%d')}_{fecha_fin.strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Error al generar Excel: {str(e)}'}), 500

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, host='0.0.0.0', port=5000) 